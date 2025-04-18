from decimal import Decimal
from django.contrib.auth.mixins import LoginRequiredMixin
from django.contrib import messages
from django.urls import reverse_lazy
from django.views.generic import ListView,TemplateView,View, DetailView,FormView, CreateView, UpdateView, DeleteView
from accountss.models import *
from room.models import *
from .mixins import *
from django.shortcuts import render, redirect,get_object_or_404
from django import forms
from .forms import *
import random
import string
from django.db.models import Count, Sum, Avg, F, ExpressionWrapper, fields
from django.db.models.functions import TruncMonth,TruncDay
from datetime import timedelta
from django.core.mail import EmailMultiAlternatives
from django.template.loader import render_to_string
from premailer import transform
import json
from django.core.serializers.json import DjangoJSONEncoder
from django.contrib.messages.views import SuccessMessageMixin
from django.template.loader import render_to_string
from datetime import datetime
import qrcode
import base64
from xhtml2pdf import pisa
from io import BytesIO
from django.shortcuts import render
from django.db.models import Count, Sum
from django.core.serializers.json import DjangoJSONEncoder
import json
from django.core.files.base import ContentFile


def admin_dashboard(request):
    # Room Type Popularity
    room_type_popularity = list(Booking.objects.values('room__room_type__name').annotate(count=Count('id')))
    room_type_popularity_data = [{'name': item['room__room_type__name'], 'y': item['count']} for item in room_type_popularity]

    # Revenue by Room Type
    revenue_by_room_type = list(Booking.objects.filter(is_paid=True).values('room__room_type__name').annotate(total_revenue=Sum('total_amount')))
    revenue_by_room_type_data = [{'name': item['room__room_type__name'], 'y': float(item['total_revenue'])} for item in revenue_by_room_type]

   
    # User Stats
    user_roles_distribution = list(Custom_user.objects.values('role').annotate(count=Count('id')))
    user_roles_distribution_data = [{'name': item['role'], 'y': item['count']} for item in user_roles_distribution]

    # User Activity Stats
    recent_users = list(Custom_user.objects.filter(last_login__isnull=False).order_by('-last_login')[:10].values('username', 'last_login'))
    recent_users_data = [{'name': item['username'], 'y': item['last_login'].timestamp() * 1000} for item in recent_users]

   # Monthly Revenue
    monthly_revenue = list(
        Booking.objects.filter(is_paid=True)
        .annotate(month=TruncMonth('created_at'))
        .values('month')
        .annotate(total_revenue=Sum('total_amount'))
        .order_by('month')
    )
    monthly_revenue_data = [{'name': item['month'].strftime('%B %Y'), 'y': float(item['total_revenue'])} for item in monthly_revenue]
    
    # Extract categories and data values separately
    monthly_revenue_categories = [item['month'].strftime('%B %Y') for item in monthly_revenue]
    monthly_revenue_values = [float(item['total_revenue']) for item in monthly_revenue]

    
    # Extract categories and data values separately
    monthly_revenue_categories = [item['month'].strftime('%B %Y') for item in monthly_revenue]
    monthly_revenue_values = [float(item['total_revenue']) for item in monthly_revenue]
    
   

    # User Registration Trends
    user_registration_trends = list(
        Custom_user.objects.annotate(month=TruncMonth('date_joined')).values('month').annotate(count=Count('id')).order_by('month')
    )
    user_registration_trends_data = [{'name': item['month'].strftime('%B %Y'), 'y': item['count']} for item in user_registration_trends]

    # Extract categories and data values separately
    user_registration_categories = [item['month'].strftime('%B %Y') for item in user_registration_trends]
    user_registration_values = [item['count'] for item in user_registration_trends]


    context = {
        'room_type_popularity_data': json.dumps(room_type_popularity_data, cls=DjangoJSONEncoder),
        'revenue_by_room_type_data': json.dumps(revenue_by_room_type_data, cls=DjangoJSONEncoder),
        'user_roles_distribution_data': json.dumps(user_roles_distribution_data, cls=DjangoJSONEncoder),
        'recent_users_data': json.dumps(recent_users_data, cls=DjangoJSONEncoder),
        'monthly_revenue_categories': json.dumps(monthly_revenue_categories),
        'monthly_revenue_values': json.dumps(monthly_revenue_values),
        'user_registration_categories': json.dumps(user_registration_categories),
        'user_registration_values': json.dumps(user_registration_values)
    }
    return render(request, 'admins/admin_dashboard.html', context)



class CustomerDetailView(OwnerRequiredMixin,DetailView):
    model = Custom_user
    template_name = 'admins/customer_detail.html'
    context_object_name = 'customer'

    def get_queryset(self):
        return Custom_user.objects.filter(role='customer')


# Detail view for Receptionist
class CustomerListView(OwnerRequiredMixin, ListView):
    model = Custom_user
    template_name = 'admins/customer_list.html'
    context_object_name = 'customers'
    paginate_by = 10

    def get_queryset(self):
        query = self.request.GET.get('search')
        if query:
            return Custom_user.objects.filter(role='customer', first_name__icontains=query)
        return Custom_user.objects.filter(role='customer').order_by('-id')


from django.core.mail import EmailMultiAlternatives
from django.template.loader import render_to_string
from django.utils.html import strip_tags
from django.contrib import messages
from django.urls import reverse_lazy
from django.views.generic import DeleteView

class CustomerDeleteView(OwnerRequiredMixin, DeleteView):
    model = Custom_user
    template_name = 'admins/customer_list.html'
    success_url = reverse_lazy('admins:customer_list')

    def post(self, request, *args, **kwargs):
        customer = self.get_object()  # Get the customer object
        email = customer.email
        full_name = customer.get_full_name()
        print(full_name)
        print(email)
        
        # Call the delete method
        response = super().delete(request, *args, **kwargs)

        # Send deletion confirmation email
        subject = 'Account Deletion Confirmation - OLIKS HOMES'
        from_email = 'olikshomes@gmail.com'

        # Render email content from template
        html_content = render_to_string('admins/customer_deletion_email_template.html', {
            'full_name': full_name,
            'support_email': 'olikshomes@gmail.com',  # You can set your support email here
        })

        # Create plain text content by stripping HTML tags
        text_content = strip_tags(html_content)

        # Create email
        email_message = EmailMultiAlternatives(
            subject=subject,
            body=text_content,
            from_email=from_email,
            to=[email]
        )

        # Attach HTML content
        email_message.attach_alternative(html_content, "text/html")
        
        # Send the email
        email_message.send()
        
        # Add the success message
        messages.success(request, f"Customer {full_name} has been deleted successfully.")

        return response

    
        

# owner views for creating manager and receptionsits
class ManagerListView(OwnerRequiredMixin, ListView):
    model = Custom_user
    template_name = 'admins/manager_list.html'
    context_object_name = 'managers'
    paginate_by = 10

    def get_queryset(self):
        query = self.request.GET.get('search')
        if query:
            return Custom_user.objects.filter(role='manager', first_name__icontains=query)
        return Custom_user.objects.filter(role='manager').order_by('-id')


class ManagerDetailView(OwnerRequiredMixin,DetailView):
    model = Custom_user
    template_name = 'admins/manager_detail.html'
    context_object_name = 'manager'

    def get_queryset(self):
        return Custom_user.objects.filter(role='manager')

class ManagerUpdateView(OwnerRequiredMixin, UpdateView):
    model = Custom_user
    form_class = CustomUserChangeForm
    template_name = 'admins/manager_form.html'
    success_url = reverse_lazy('admins:manager_list')

class ManagerDeleteView(OwnerRequiredMixin, DeleteView):
    model = Custom_user
    template_name = 'admins/manager_list.html'
    success_url = reverse_lazy('admins:manager_list')
    def post(self, request, *args, **kwargs):
        manager = self.get_object()
        full_name = manager.get_full_name()
        response = super().delete(request, *args, **kwargs)
        messages.success(request, f"Manager {full_name} has been deleted successfully.")

        return response
  
class ReceptionistListView(OwnerOrManagerRequiredMixin, ListView):
    model = Custom_user
    template_name = 'admins/receptionist_list.html'
    context_object_name = 'receptionists'
    paginate_by = 10

    def get_queryset(self):
        query = self.request.GET.get('search')
        if query:
            return Custom_user.objects.filter(role='receptionist', first_name__icontains=query)
        return Custom_user.objects.filter(role='receptionist').order_by('-id')


class ManagerCreateView(OwnerRequiredMixin, CreateView):
    model = Custom_user
    form_class = CustomUserCreationForm
    template_name = 'admins/manager_form.html'
    success_url = reverse_lazy('admins:manager_list')

    def form_valid(self, form):
        form.instance.role = 'manager'
        response = super().form_valid(form)

        # Email confirmation logic
        manager = form.instance
        emails = form.cleaned_data.get('email')
        password = form.cleaned_data.get('password1')  # Assuming this is the password field name
        print(password)


        
        # URL for login
        login_url = f"{BASE_URL}/login/"

        # Render email content from template
        html_content = render_to_string('admins/manager_account_creation_confirmation_email_template.html', {
            'manager': manager,
            'password': password,
            'login_url': login_url,
        })
        
        # Inline CSS
        html_content = transform(html_content)
        print(emails)
        # Create the email message
        email = EmailMultiAlternatives(
            subject='Manager Account Created - Please Change Your Password',
            from_email='olikshomes@gmail.com',
            to=[emails]
        )
        
        # Attach the HTML content
        email.attach_alternative(html_content, "text/html")
        
        # Send the email
        email.send()

        return response


class ReceptionistUpdateView(OwnerOrManagerRequiredMixin, UpdateView):
    model = Custom_user
    form_class = CustomUserChangeForm
    template_name = 'admins/receptionist_form.html'
    success_url = reverse_lazy('admins:receptionist_list')

class ReceptionistDetailView(OwnerOrManagerRequiredMixin,DetailView):
    model = Custom_user
    template_name = 'admins/receptionist_detail.html'
    context_object_name = 'receptionist'

    def get_queryset(self):
        return Custom_user.objects.filter(role='receptionist')



class ReceptionistCreateView(OwnerOrManagerRequiredMixin, CreateView):
    model = Custom_user
    form_class = CustomUserCreationForm
    template_name = 'admins/receptionist_form.html'
    success_url = reverse_lazy('admins:receptionist_list')

    def form_valid(self, form):
        form.instance.role = 'receptionist'
        response = super().form_valid(form)

        # Email confirmation logic
        receptionist = form.instance
        password = form.cleaned_data.get('password1')  # Assuming this is the password field name
        receptionist_email = form.cleaned_data.get('email')  # Assuming this is the password field name
        # URL for login
        login_url = f"{BASE_URL}/login/"

        # Render email content from template
        html_content = render_to_string('admins/receptionist_account_creation_confirmation_email_template.html', {
            'receptionist': receptionist,
            'password': password,
            'login_url': login_url,
        })
        
        # Inline CSS
        html_content = transform(html_content)

        # Create the email message
        email = EmailMultiAlternatives(
            subject='Receptionist Account Created - Please Change Your Password',
            from_email='olikshomes@gmail.com',
            to=[receptionist_email]
        )
        
        # Attach the HTML content
        email.attach_alternative(html_content, "text/html")
        
        # Send the email
        email.send()

        return response



class ReceptionistDeleteView(OwnerOrManagerRequiredMixin, DeleteView):
    model = Custom_user
    template_name = 'admins/receptionist_list.html'
    success_url = reverse_lazy('admins:receptionist_list')
    def post(self, request, *args, **kwargs):
        receptionist = self.get_object()
        full_name = receptionist.get_full_name()
        response = super().delete(request, *args, **kwargs)
        messages.success(request, f"Receptionist {full_name} has been deleted successfully.")

        return response



# Category Views
class CategoryListView(OwnerManagerOrReceptionistRequiredMixin,ListView):
    model = Category
    template_name = 'admins/category_list.html'
    context_object_name = 'categories'
    paginate_by = 10

    def get_queryset(self):
        query = self.request.GET.get('search')
        if query:
            return Category.objects.filter(name__icontains=query)
        return Category.objects.order_by('name')

class CategoryDetailView(OwnerManagerOrReceptionistRequiredMixin, DetailView):
    model = Category
    template_name = 'admins/category_detail.html'



class CategoryCreateView(OwnerOrManagerRequiredMixin,  CreateView,SuccessMessageMixin):
    model = Category
    template_name = 'admins/category_form.html'
    form_class = CategoryForm
    success_url = reverse_lazy('admins:category_list')
    success_message = "Room category was created successfully."

    
class CategoryUpdateView(OwnerOrManagerRequiredMixin, UpdateView):
    model = Category
    template_name = 'admins/category_form.html'
    form_class = CategoryForm
    success_url = reverse_lazy('admins:category_list')
    success_message = "Room category was updated successfully."


class CategoryDeleteView(OwnerOrManagerRequiredMixin, DeleteView):
    model = Category
    template_name = 'admins/category_confirm_delete.html'
    success_url = reverse_lazy('admins:category_list')

# Room Views
class RoomListView(OwnerManagerOrReceptionistRequiredMixin, ListView):
    model = Room
    template_name = 'admins/room_list.html'
    context_object_name = 'object_list'
    paginate_by = 10

    def get_queryset(self):
        queryset = Room.objects.all().order_by('id')
        
        # Annotate the queryset with the average rating
        queryset = queryset.annotate(average_rating=Avg('roomrating__rating'))
        
        search_query = self.request.GET.get('search', '')
        if search_query:
            queryset = queryset.filter(
                Q(room_number__icontains=search_query) | 
                Q(room_type__name__icontains=search_query)
            )
        return queryset


class RoomDetailView(OwnerManagerOrReceptionistRequiredMixin, DetailView):
    model = Room
    template_name = 'admins/room_detail.html'

class RoomCreateView(OwnerOrManagerRequiredMixin, CreateView):
    model = Room
    template_name = 'admins/room_form.html'
    form_class = RoomForm
    success_url = reverse_lazy('admins:room_list')

    def form_valid(self, form):
        response = super().form_valid(form)
        messages.success(self.request, 'Room created successfully.')
        return response

class RoomUpdateView(OwnerOrManagerRequiredMixin, UpdateView):
    model = Room
    template_name = 'admins/room_form.html'
    form_class = RoomForm
    success_url = reverse_lazy('admins:room_list')

    def form_valid(self, form):
        response = super().form_valid(form)
        messages.success(self.request, 'Room updated successfully.')
        return response

class RoomDeleteView(OwnerOrManagerRequiredMixin,  DeleteView):
    model = Room
    template_name = 'admins/room_confirm_delete.html'
    success_url = reverse_lazy('admins:room_list')




# Booking Views
from django.views.generic import ListView
from django.db.models import Q
from room.models import Booking
from datetime import timedelta
from django.utils import timezone

class BookingListView(OwnerManagerOrReceptionistRequiredMixin, ListView):
    model = Booking
    template_name = 'admins/booking_list.html'
    context_object_name = 'object_list'
    paginate_by = 10

    def get_queryset(self):
        queryset = Booking.objects.all().order_by('-created_at')

        # Get the current time and two days ago
        now = timezone.now()
        two_days_ago = now - timedelta(days=2)

        # Fetch bookings that have a past checkout date and are either pending or confirmed
        past_bookings = queryset.filter(check_out_date__lt=now, status__in=['pending', 'confirmed'])

        for booking in past_bookings:
            # Update booking status to 'cancelled'
            booking.status = 'cancelled'

            # Update room status and booking details
            booking.room.room_status = 'vacant'
            booking.checked_in = False
            booking.checked_out = True

            # Save the updates to room and booking
            booking.room.save()
            booking.save()

        # Automatically cancel bookings that are pending and created more than two days ago
        queryset.filter(created_at__lt=two_days_ago, status='pending').update(status='cancelled')

        # Reapply ordering by '-created_at'
        queryset = Booking.objects.all().order_by('-created_at')

        # Handle search functionality
        search_query = self.request.GET.get('search', '')
        if search_query:
            queryset = queryset.filter(
                Q(room__room_number__icontains=search_query) |
                Q(room__room_type__name__icontains=search_query) |
                Q(full_name__icontains=search_query) |
                Q(tx_ref__icontains=search_query) |
                Q(status__icontains=search_query)
            )

        return queryset


class BookingVerifyView(View):
    def get(self, request, booking_id, *args, **kwargs):
        try:
            booking = Booking.objects.get(id=booking_id)
        except Booking.DoesNotExist:
            return render(request, 'admins/booking_not_found.html')
        
        return render(request, 'admins/booking_verify.html', {'booking': booking})
class BookingDetailView(OwnerManagerOrReceptionistRequiredMixin, DetailView):
    model = Booking
    template_name = 'admins/booking_detail.html'

class BookingCreateView(OwnerManagerOrReceptionistRequiredMixin, CreateView):
    model = Booking
    form_class = BookingCreateForm
    template_name = 'admins/booking_form.html'

    def form_valid(self, form):
        booking = form.save(commit=False)
        full_name = form.cleaned_data['full_name']
        booking.tx_ref = f"booking-{full_name.replace(' ', '')}-tx-{''.join(random.choices(string.ascii_lowercase + string.digits, k=10))}"
        booking.status = 'pending'
        
        if booking.check_in_date and booking.check_out_date:
            duration = (booking.check_out_date - booking.check_in_date).days
            booking.original_booking_amount = booking.room.price_per_night * duration
            booking.total_amount = booking.original_booking_amount
        booking.save()
        return redirect('admins:payment_create',booking_id=booking.id)



class BookingUpdateView(OwnerManagerOrReceptionistRequiredMixin, UpdateView):
    model = Booking
    template_name = 'admins/booking_update.html'
    form_class = BookingUpdateForm
    success_url = reverse_lazy('admins:booking_list')

    def form_valid(self, form):
        response = super().form_valid(form)
        booking = form.instance
        return response

from django.http import HttpResponse, Http404

class DownloadIDImageView(View):
    def get(self, request, pk):
        try:
            booking = Booking.objects.get(pk=pk)
            if booking.id_image:
                response = HttpResponse(booking.id_image, content_type='image/jpeg')
                response['Content-Disposition'] = f'attachment; filename="{booking.id_image.name}"'
                return response
            else:
                raise Http404("No ID image available for this booking.")
        except Booking.DoesNotExist:
            raise Http404("Booking not found.")





from django.urls import reverse

class BookingExtendView(OwnerManagerOrReceptionistRequiredMixin, UpdateView):
    model = Booking
    form_class = BookingExtendForm
    template_name = 'admins/booking_extend_form.html'

    def form_valid(self, form):
        booking = form.save(commit=False)
        additional_amount = booking.calculate_additional_amount()
        booking.booking_extend_amount = additional_amount
        booking.total_amount += additional_amount
        booking.status = 'pending'
        
        # Save the booking first
        booking.save()
        
        # Create or update the Payment object
        payment, created = Payment.objects.get_or_create(
            booking=booking,
            
        )
        
        if not created:
            # If the payment already exists, update it
            payment.save()
        
        # Generate a new tx_ref for the payment
        full_name = booking.full_name
        booking.tx_ref = f"booking-{full_name.replace(' ', '')}-tx-{''.join(random.choices(string.ascii_lowercase + string.digits, k=10))}"
        booking.save()
        
        # Redirect to the payment extension view with booking and payment IDs
        return redirect(reverse('admins:payment_extend_update', kwargs={'booking_id': booking.id, 'pk': payment.id}))


# Room Payment Views


from django.utils.safestring import mark_safe

class PaymentExtendView(OwnerManagerOrReceptionistRequiredMixin, UpdateView):
    model = Payment
    form_class = PaymentExtendForm
    template_name = 'admins/payment_extend_form.html'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        booking_id = self.kwargs.get('booking_id')
        context['booking'] = get_object_or_404(Booking, pk=booking_id)
        return context

    def get_object(self):
        booking_id = self.kwargs.get('booking_id')
        payment_id = self.kwargs.get('pk')
        booking = get_object_or_404(Booking, pk=booking_id)
        return get_object_or_404(Payment, id=payment_id, booking=booking)

    def form_valid(self, form):
        payment = form.save(commit=False)
        booking_id = self.kwargs.get('booking_id')
        booking = get_object_or_404(Booking, pk=booking_id)
        payment.booking = booking
        payment.transaction_id = booking.tx_ref
        payment.payment_date = datetime.now()
        payment.status = 'completed'
        payment.save()
        booking.status = 'confirmed'
        booking.check_out_date = booking.extended_check_out_date
        booking.save()

        # Generate receipt PDF
        pdf_response = self.generate_pdf(booking)
        pdf_name = f"room_booking_extend_receipt_{booking.id}_{booking.full_name if booking.full_name else booking.user.username}.pdf"
        pdf_file = ContentFile(pdf_response) 
        payment.receipt_pdf.save(pdf_name, pdf_file)
        if booking.email2:
                    from_email = 'olikshomes@gmail.com'
                    subject = 'Room Booking Extend Confirmation'
                    html_content = render_to_string('checkout_date_extenstion_email_template.html', {'booking': booking})
                    # Inline CSS
                    html_content = transform(html_content)

                    # Render email content from template
                    
                    user_email = booking.email2 
                    
                    # Create email
                    email_message = EmailMultiAlternatives(
                        subject=subject,
                        body=html_content,
                        from_email=from_email,
                        to=[user_email]
                    )

                    # Attach HTML content
                    email_message.attach_alternative(html_content, "text/html")
                    email_message.attach(f'receipt_{booking.id}_{booking.full_name}.pdf', pdf_response, 'application/pdf')
                    
                    # Send the email
                    email_message.send()

        if self.request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return HttpResponse(pdf_response, content_type='application/pdf')

        # Automatically download the PDF receipt
        response = HttpResponse(pdf_response, content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="receipt_{booking.id}_{booking.full_name}.pdf"'

        messages.success(self.request, 'Booking and Payment for Extension completed')
        return response

    def generate_pdf(self, booking):
        buffer = BytesIO()
        
        # Generate QR code data
        qr_code_data = self.generate_qr_code(f'{BASE_URL}/admins/verify/{booking.id}')
        
        context = {
            'booking': booking,
            'qr_code_data': qr_code_data,
        }
        
        html_string = render_to_string('room/checkout_date_extenstion_email_template_receipt.html', context)
        
        pisa_status = pisa.CreatePDF(html_string, dest=buffer)
        buffer.seek(0)
        return buffer.read()

    def generate_qr_code(self, data):
        qr = qrcode.QRCode(
            version=1,
            error_correction=qrcode.constants.ERROR_CORRECT_L,
            box_size=10,
            border=4,
        )
        qr.add_data(data)
        qr.make(fit=True)
        img = qr.make_image(fill_color="black", back_color="white")
        buffered = BytesIO()
        img.save(buffered, format="PNG")
        img_str = base64.b64encode(buffered.getvalue()).decode()
        return mark_safe(f'data:image/png;base64,{img_str}')

from django.db import IntegrityError, transaction
class PaymentCreateView(OwnerManagerOrReceptionistRequiredMixin, CreateView):
    model = Payment
    form_class = PaymentCreateForm
    template_name = 'admins/payment_form.html'
    success_url = reverse_lazy('admins:payment_list')

    def get_context_data(self, **kwargs):
        context = super(PaymentCreateView, self).get_context_data(**kwargs)
        booking_id = self.kwargs.get('booking_id')  
        context['booking'] = get_object_or_404(Booking, pk=booking_id)
        return context

    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        booking_id = self.kwargs.get('booking_id')
        booking = get_object_or_404(Booking, pk=booking_id)
        kwargs.update({'initial': {'booking': booking}})
        return kwargs

    def form_valid(self, form):
        booking_id = self.kwargs.get('booking_id')
        booking = get_object_or_404(Booking, pk=booking_id)

        # Check if a payment already exists
        if Payment.objects.filter(booking=booking).exists():
            messages.error(self.request, 'Payment already exists for this booking.')
            return redirect('admins:payment_list')

        try:
            with transaction.atomic():
                payment = form.save(commit=False)
                payment.booking = booking
                payment.transaction_id = booking.tx_ref
                payment.payment_date = datetime.now()
                payment.status = 'completed'
                payment.save()

                booking.status = 'confirmed'
                booking.save()

                # Generate receipt PDF
                pdf_response = self.generate_pdf(booking)
                pdf_name = f"room_booking_receipt_{booking.id}_{booking.full_name if booking.full_name else booking.user.username}.pdf"
                pdf_file = ContentFile(pdf_response) 
                payment.receipt_pdf.save(pdf_name, pdf_file)
                if booking.email2:
                    from_email = 'olikshomes@gmail.com'
                    subject = 'Room Booking Confirmation'
                    html_content = render_to_string('room/booking_confirmation_template.html', {'booking': booking})
                    # Inline CSS
                    html_content = transform(html_content)

                    # Render email content from template
                    
                    user_email = booking.email2 
                    
                    # Create email
                    email_message = EmailMultiAlternatives(
                        subject=subject,
                        body=html_content,
                        from_email=from_email,
                        to=[user_email]
                    )

                    # Attach HTML content
                    email_message.attach_alternative(html_content, "text/html")
                    email_message.attach(f'receipt_{booking.id}_{booking.full_name}.pdf', pdf_response, 'application/pdf')
                    
                    # Send the email
                    email_message.send()

                # Handle AJAX request for PDF download
                if self.request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                    return HttpResponse(pdf_response, content_type='application/pdf')

                # Regular response (fallback)
                response = HttpResponse(pdf_response, content_type='application/pdf')
                response['Content-Disposition'] = f'attachment; filename="receipt_{booking.id}_{booking.full_name}.pdf"'
                
            
                
                
                
                messages.success(self.request, 'Booking and Payment completed')
                return response

        except IntegrityError:
            messages.error(self.request, 'A payment for this booking already exists.')
            return redirect('admins:payment_list')


    def generate_pdf(self, booking):
        buffer = BytesIO()
        
        # Generate QR code data
        qr_code_data = self.generate_qr_code(f'{BASE_URL}/admins/verify/{booking.id}')
        
        context = {
            'booking': booking,
            'qr_code_data': qr_code_data,
        }
        
        
        html_string = render_to_string('room/booking_confirmation_template_receipt.html', context)
        
        pisa_status = pisa.CreatePDF(html_string, dest=buffer)
        buffer.seek(0)
        return buffer.read()

    def generate_qr_code(self, data):
        qr = qrcode.QRCode(
            version=1,
            error_correction=qrcode.constants.ERROR_CORRECT_L,
            box_size=10,
            border=4,
        )
        qr.add_data(data)
        qr.make(fit=True)
        img = qr.make_image(fill_color="black", back_color="white")
        buffered = BytesIO()
        img.save(buffered, format="PNG")
        img_str = base64.b64encode(buffered.getvalue()).decode()
        return mark_safe(f'data:image/png;base64,{img_str}')



class PaymentListView(OwnerManagerOrReceptionistRequiredMixin, ListView):
    model = Payment
    template_name = 'admins/payment_list.html'
    context_object_name = 'object_list'
    paginate_by = 10

    def get_queryset(self):
        queryset = Payment.objects.all().order_by('-payment_date')
        search_query = self.request.GET.get('search', '')
        if search_query:
            queryset = queryset.filter(
                Q(booking__room__room_number__icontains=search_query) |
                Q(booking__room__room_type__name__icontains=search_query) |
                Q(booking__user__username__icontains=search_query) |
                Q(transaction_id__icontains=search_query)
            )
        return queryset



class PaymentDetailView(OwnerManagerOrReceptionistRequiredMixin, DetailView):
    model = Payment
    template_name = 'admins/payment_detail.html'


from django.http import HttpResponseRedirect
class PaymentDeleteView(OwnerOrManagerRequiredMixin, DeleteView):
    model = Payment
    template_name = 'admins/payment_confirm_delete.html'
    success_url = reverse_lazy('admins:payment_list')

    def delete(self, request, *args, **kwargs):
        self.object = self.get_object()
        booking = self.object.booking  # Get the associated booking instance

        if booking:
            print(f"Booking found: {booking.id}")  # Debug statement
        else:
            print("No associated booking found.")  # Debug statement
        
        success_url = self.get_success_url()

        # Explicitly delete the associated booking first, then the payment
        if booking:
            booking.delete()
        
        self.object.delete()

        return HttpResponseRedirect(success_url)


class PaymentDownloadReceiptView(View):
    def get(self, request, pk, *args, **kwargs):
        payment = get_object_or_404(Payment, pk=pk)

        # Serve the file for download
        response = HttpResponse(payment.receipt_pdf, content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="{payment.receipt_pdf.name}"'
        return response



class RoomReportView(View):
    def get(self, request):
        current_month = now().month
        current_year = now().year

        # Monthly room bookings
        monthly_room_bookings = Booking.objects.filter(
            created_at__year=current_year, created_at__month=current_month
        ).count()

        # Monthly room revenue
        monthly_room_revenue = Payment.objects.filter(
            payment_date__year=current_year, payment_date__month=current_month
        ).aggregate(total=Sum('booking__total_amount'))['total'] or 0

        # Room popularity
        popular_rooms = Booking.objects.values('room__room_number').annotate(total=Count('id')).order_by('-total')

        # Guest satisfaction
        room_ratings = RoomRating.objects.values('room__room_number').annotate(avg_rating=Avg('rating')).order_by('-avg_rating')

        context = {
            'monthly_room_bookings': monthly_room_bookings,
            'monthly_room_revenue': monthly_room_revenue,
            'popular_rooms': popular_rooms,
            'room_ratings': room_ratings,
        }
        
        return render(request, 'admins/room_reports.html', context)

from openpyxl import Workbook
from datetime import datetime


from datetime import datetime, timedelta



class ExportRoomReportView(View):

    def post(self, request, report_type):
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename={report_type}_report.xlsx'

        wb = Workbook()
        ws = wb.active
        ws.title = f'{report_type.capitalize()} Report'

        start_month = request.POST.get('start_month')
        end_month = request.POST.get('end_month')

        if report_type == 'bookings':
            bookings = Booking.objects.all()
            ws.append(['Booking ID', 'Room Number', 'User', 'Check-in Date', 'Check-out Date', 'Total Amount', 'Status'])
            for booking in bookings:
                user_display = booking.full_name #.user.username if booking.user else booking.full_name
                ws.append([
                    booking.id,
                    booking.room.room_number,
                    user_display,
                    booking.check_in_date,
                    booking.check_out_date,
                    booking.total_amount,
                    booking.status,
                ])

        elif report_type == 'monthly_revenue':
            current_month = now().month
            current_year = now().year
            payments = Payment.objects.filter(payment_date__year=current_year, payment_date__month=current_month)
            ws.append(['Payment ID', 'Booking ID', 'Total Amount', 'Payment Date', 'Payment Method'])

            total_amount = 0
            for payment in payments:
                total_amount += payment.booking.toal_amount
                ws.append([
                    payment.id,
                    payment.booking.id,
                    payment.booking.toal_amount,
                    payment.payment_date,
                    payment.payment_method,
                ])
            ws.append(['', '', 'Total', total_amount])

        elif report_type == 'popularity':
            popular_rooms = Booking.objects.values('room__room_number').annotate(total=Count('id')).order_by('-total')
            ws.append(['Room Number', 'Total Bookings'])
            for room in popular_rooms:
                ws.append([room['room__room_number'], room['total']])

        elif report_type == 'satisfaction':
            room_ratings = RoomRating.objects.values('room__room_number').annotate(avg_rating=Avg('rating')).order_by('-avg_rating')
            ws.append(['Room Number', 'Average Rating'])
            for rating in room_ratings:
                ws.append([rating['room__room_number'], rating['avg_rating']])

        elif report_type == 'all_payments':
            payments = Payment.objects.all()
            ws.append(['Payment ID', 'Booking ID', 'Total Amount', 'Payment Date', 'Payment Method'])
            for payment in payments:
                # Convert payment_date to naive datetime
                payment_date_naive = payment.payment_date.replace(tzinfo=None)
                ws.append([
                    payment.id,
                    payment.booking.id,
                    payment.booking.toal_amount,
                    payment_date_naive,
                    payment.payment_method,
                ])

        # POST request with date range (monthly reports)
        if start_month and end_month:
            try:
                start_date = datetime.strptime(start_month, '%Y-%m').date()
                end_date = datetime.strptime(end_month, '%Y-%m').date()
            except ValueError:
                return HttpResponse("Invalid date format", status=400)

            if report_type == 'monthly_bookings':
                current_date = start_date
                bookings_data = []
                month_totals = {}

                while current_date <= end_date:
                    next_month = current_date.replace(day=28) + timedelta(days=4)  # this will get to the next month
                    next_month_start = next_month - timedelta(days=next_month.day - 1)

                    bookings = Booking.objects.filter(
                        check_in_date__gte=current_date,
                        check_in_date__lt=next_month_start
                    )

                    monthly_count = bookings.count()
                    month_totals[current_date.strftime('%Y-%m')] = monthly_count

                    for booking in bookings:
                        user_display = booking.user.username if booking.user else booking.full_name
                        bookings_data.append([
                            booking.id,
                            booking.created_at.replace(tzinfo=None) if booking.created_at else None,
                            booking.room.room_number,
                            user_display,
                            booking.check_in_date,
                            booking.check_out_date,
                            booking.total_amount,
                            booking.status
                        ])

                    current_date = next_month_start  # move to the next month

                # Write booking details
                ws.append(['Booking ID','Booking Date', 'Room Number', 'User', 'Check-in Date', 'Check-out Date', 'Total Amount', 'Status'])
                for row in bookings_data:
                    ws.append(row)

                # Write summary of monthly bookings
                ws.append([])
                ws.append(['Month', 'Total Bookings'])
                for month, total in month_totals.items():
                    ws.append([month, total])

            elif report_type == 'monthly_revenue':
                current_date = start_date
                revenue_data = []
                month_totals = {}

                while current_date <= end_date:
                    next_month = current_date.replace(day=28) + timedelta(days=4)  # this will get to the next month
                    next_month_start = next_month - timedelta(days=next_month.day - 1)

                    payments = Payment.objects.filter(
                        payment_date__gte=current_date,
                        payment_date__lt=next_month_start
                    )

                    monthly_total = payments.aggregate(total=Sum('booking__total_amount'))['total'] or 0
                    month_totals[current_date.strftime('%Y-%m')] = monthly_total

                    for payment in payments:
                        payment_date_naive = payment.payment_date.replace(tzinfo=None)  # Convert to naive datetime
                        revenue_data.append([
                            payment.id,
                            payment.booking.id,
                            payment.booking.total_amount,
                            payment_date_naive,
                            payment.payment_method
                        ])

                    current_date = next_month_start  # move to the next month

                # Write payment details
                ws.append(['Payment ID', 'Booking ID', 'Total Amount', 'Payment Date', 'Payment Method'])
                for row in revenue_data:
                    ws.append(row)

                # Write summary of monthly revenue
                ws.append([])
                ws.append(['Month', 'Total Revenue'])
                for month, total in month_totals.items():
                    ws.append([month, total])

        wb.save(response)
        return response





class HallReportView(View):
    def get(self, request):
        current_month = now().month
        current_year = now().year

        # Monthly hall bookings
        monthly_hall_bookings = Hall_Booking.objects.filter(
            created_at__year=current_year, created_at__month=current_month
        ).count()

        # Monthly hall revenue
        monthly_hall_revenue = Hall_Booking.objects.filter(
            created_at__year=current_year, created_at__month=current_month
        ).aggregate(total=Sum('amount_due'))['total'] or 0

        # Hall popularity (number of bookings)
        popular_halls = Hall_Booking.objects.values('hall__hall_number').annotate(total=Count('id')).order_by('-total')

        # Guest satisfaction (average rating - assuming there's a HallRating model similar to RoomRating)
        # hall_ratings = HallRating.objects.values('hall__hall_number').annotate(avg_rating=Avg('rating')).order_by('-avg_rating')

        context = {
            'monthly_hall_bookings': monthly_hall_bookings,
            'monthly_hall_revenue': monthly_hall_revenue,
            'popular_halls': popular_halls,
            # 'hall_ratings': hall_ratings,  # Uncomment if ratings are implemented
        }
        
        return render(request, 'admins/hall_reports.html', context)


class ExportHallReportView(View):
    def post(self, request, report_type):
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename={report_type}_report.xlsx'

        wb = Workbook()
        ws = wb.active
        ws.title = f'{report_type.capitalize()} Report'

        start_month = request.POST.get('start_month')
        end_month = request.POST.get('end_month')

        if start_month and end_month:
            start_date = datetime.strptime(start_month, '%Y-%m')
            end_date = datetime.strptime(end_month, '%Y-%m')

        if report_type == 'bookings':
            bookings = Hall_Booking.objects.all()
            ws.append(['Booking ID', 'Hall Number', 'User', 'Start Date', 'End Date','Start Time','End Time', 'Total Amount', 'Status'])
            for booking in bookings:
                user_display = booking.user.username if booking.user else booking.full_name
                ws.append([
                    booking.id,
                    booking.hall.hall_number,
                    user_display,
                    booking.start_date,
                    booking.end_date,
                    booking.start_time,
                    booking.end_time,
                    booking.amount_due,
                    booking.status,
                ])

        elif report_type == 'all_payments':
            payments = Hall_Payment.objects.all()
            ws.append(['Payment ID', 'Booking ID', 'Payment Method', 'Transaction ID', 'Payment Date', 'Status'])
            for payment in payments:
                payment_date_naive = payment.payment_date.replace(tzinfo=None)
                ws.append([
                    payment.id,
                    payment.booking.id,
                    payment.payment_method,
                    payment.transaction_id,
                    payment_date_naive,
                    payment.status,
                ])

        # Monthly Payments Report
        elif report_type == 'monthly_payments':
            current_date = start_date
            payment_data = []
            month_totals = {}

            # Loop through each month in the date range
            while current_date <= end_date:
                next_month = current_date.replace(day=28) + timedelta(days=4)
                next_month_start = next_month - timedelta(days=next_month.day - 1)

                # Get payments for the current month
                payments = Hall_Payment.objects.filter(
                    payment_date__gte=current_date,
                    payment_date__lt=next_month_start
                )

                # Calculate the total for the current month
                monthly_total = payments.aggregate(total=Sum('booking__amount_due'))['total'] or 0
                month_totals[current_date.strftime('%Y-%m')] = monthly_total

                # Collect payment details
                for payment in payments:
                    payment_date_naive = payment.payment_date.replace(tzinfo=None)  # Convert to naive datetime
                    payment_data.append([
                        payment.id,
                        payment.booking.id,
                        payment.booking.amount_due,
                        payment_date_naive,
                        payment.payment_method
                    ])

                current_date = next_month_start  # Move to the next month

            # Write payment details to Excel
            ws.append(['Payment ID', 'Booking ID', 'Total Amount', 'Payment Date', 'Payment Method'])
            for row in payment_data:
                ws.append(row)

            # Add a space before the summary section
            ws.append([])
            ws.append([])

            # Write monthly totals to Excel
            ws.append(['Month', 'Total Payments'])
            for month, total in month_totals.items():
                ws.append([month, total])

        # Monthly Revenue Report
        elif report_type == 'monthly_revenue':
            current_date = start_date
            revenue_data = []
            month_totals = {}

            while current_date <= end_date:
                next_month = current_date.replace(day=28) + timedelta(days=4)
                next_month_start = next_month - timedelta(days=next_month.day - 1)

                payments = Hall_Payment.objects.filter(
                    payment_date__gte=current_date,
                    payment_date__lt=next_month_start
                )

                monthly_total = payments.aggregate(total=Sum('booking__amount_due'))['total'] or 0
                month_totals[current_date.strftime('%Y-%m')] = monthly_total

                for payment in payments:
                    payment_date_naive = payment.payment_date.replace(tzinfo=None)
                    revenue_data.append([
                        payment.id,
                        payment.booking.id,
                        payment.booking.amount_due,
                        payment_date_naive,
                        payment.payment_method
                    ])

                current_date = next_month_start

            # Write payment details
            ws.append(['Payment ID', 'Booking ID', 'Total Amount', 'Payment Date', 'Payment Method'])
            for row in revenue_data:
                ws.append(row)

            # Write summary of monthly revenue
            ws.append([])
            ws.append(['Month', 'Total Revenue'])
            for month, total in month_totals.items():
                ws.append([month, total])

        elif report_type == 'monthly_bookings':
            current_date = start_date
            bookings_data = []
            month_totals = {}

            while current_date <= end_date:
                next_month = current_date.replace(day=28) + timedelta(days=4)
                next_month_start = next_month - timedelta(days=next_month.day - 1)

                bookings = Hall_Booking.objects.filter(
                    start_date__gte=current_date,
                    start_date__lt=next_month_start
                )

                monthly_total = bookings.aggregate(total=Count('id'))['total'] or 0
                month_totals[current_date.strftime('%Y-%m')] = monthly_total
                
                for booking in bookings:
                    created_date_naive = booking.created_at.replace(tzinfo=None)
                    user_display = booking.user.username if booking.user else booking.full_name
                    bookings_data.append([
                        booking.id,
                        created_date_naive,
                        booking.hall.hall_number,
                        user_display,
                        booking.start_date,
                        booking.end_date,
                        booking.start_time,
                        booking.end_time,
                        booking.amount_due,
                        booking.status
                    ])

                current_date = next_month_start

            # Write booking details
            ws.append(['Booking ID','Created', 'Hall Number', 'User', 'Start Date', 'End Date','Start Time','End Time', 'Total Amount', 'Status'])
            for row in bookings_data:
                ws.append(row)

            # Write summary of monthly bookings
            ws.append([])
            ws.append(['Month', 'Total Bookings'])
            for month, total in month_totals.items():
                ws.append([month, total])
        # Hall Popularity Report
        elif report_type == 'popularity':
            hall_popularity = Hall_Booking.objects.values('hall__hall_number').annotate(total=Count('id')).order_by('-total')
            ws.append(['Hall Number', 'Total Bookings'])
            for hall in hall_popularity:
                ws.append([hall['hall__hall_number'], hall['total']])

        
        wb.save(response)
        return response



class MembershipReportView(View):
    def get(self, request):
        memberships = Membership.objects.all()
        membership_plans = MembershipPlan.objects.all()
        payments = MembershipPayment.objects.all()

        current_month = now().month
        current_year = now().year

        # Calculated reports
        monthly_memberships = memberships.filter(created_at__year=current_year, created_at__month=current_month).count()
        monthly_revenue = payments.filter(payment_date__year=current_year, payment_date__month=current_month).aggregate(total=Sum('amount'))['total'] or 0

        # Popularity of plans
        memberships_by_plan = memberships.values('plan__name').annotate(total=Count('id')).order_by('-total')

        context = {
            'memberships': memberships,
            'membership_plans': membership_plans,
            'payments': payments,
            'monthly_memberships': monthly_memberships,
            'monthly_revenue': monthly_revenue,
            'memberships_by_plan': memberships_by_plan,
        }

        return render(request, 'admins/membership_reports.html', context)

class ExportMembershipReportView(View):
    
    def post(self, request, report_type):
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename={report_type}_report.xlsx'

        wb = Workbook()
        ws = wb.active
        ws.title = f'{report_type.capitalize()} Report'

        start_month = request.POST.get('start_month')
        end_month = request.POST.get('end_month')

        if start_month and end_month:
            try:
                start_date = datetime.strptime(start_month, '%Y-%m').date()
                end_date = datetime.strptime(end_month, '%Y-%m').date()
            except ValueError:
                return HttpResponse("Invalid date format", status=400)

        if report_type == 'memberships':
            memberships = Membership.objects.all()
            ws.append(['Membership ID', 'User', 'Plan', 'Start Date', 'End Date', 'Status', 'Created At'])
            for membership in memberships:
                ws.append([
                    membership.id,
                    membership.user.username if membership.user else f"{membership.for_first_name} {membership.for_last_name}",
                    membership.plan.name,
                    membership.start_date.strftime('%Y-%m-%d'),
                    membership.end_date.strftime('%Y-%m-%d'),
                    membership.status,
                    membership.created_at.strftime('%Y-%m-%d %H:%M:%S')
                ])

        elif report_type == 'plans':
            plans = MembershipPlan.objects.all()
            ws.append(['Plan ID', 'Name', 'Price', 'Duration (Months)', 'Description'])
            for plan in plans:
                ws.append([plan.id, plan.name, plan.price, plan.duration_months, plan.description])

        elif report_type == 'payments':
            payments = MembershipPayment.objects.all()
            ws.append(['Payment ID', 'Membership ID', 'User', 'Amount', 'Payment Date', 'Payment Method', 'Status'])
            for payment in payments:
                user_display = payment.membership.user.username if payment.membership.user else f"{payment.membership.for_first_name} {payment.membership.for_last_name}"
                ws.append([
                    payment.id,
                    payment.membership.id,
                    user_display,
                    payment.amount,
                    payment.payment_date.strftime('%Y-%m-%d %H:%M:%S'),
                    payment.payment_method,
                    payment.status
                ])

        elif report_type == 'monthly_memberships':
            current_date = start_date
            memberships_data = []
            month_totals = {}

            while current_date <= end_date:
                next_month = current_date.replace(day=28) + timedelta(days=4)
                next_month_start = next_month - timedelta(days=next_month.day - 1)

                monthly_memberships = Membership.objects.filter(
                    created_at__gte=current_date,
                    created_at__lt=next_month_start
                )

                monthly_count = monthly_memberships.count()
                month_totals[current_date.strftime('%Y-%m')] = monthly_count

                for membership in monthly_memberships:
                    memberships_data.append([
                        membership.id,
                        membership.user.username if membership.user else f"{membership.for_first_name} {membership.for_last_name}",
                        membership.plan.name,
                        membership.start_date.strftime('%Y-%m-%d'),
                        membership.created_at.strftime('%Y-%m-%d')
                    ])

                current_date = next_month_start

            ws.append(['Membership ID', 'User', 'Plan', 'Start Date', 'Created At'])
            for row in memberships_data:
                ws.append(row)

            ws.append([])
            ws.append(['Month', 'Total Memberships'])
            for month, total in month_totals.items():
                ws.append([month, total])

        elif report_type == 'monthly_revenue':
            current_date = start_date
            revenue_data = []
            month_totals = {}

            while current_date <= end_date:
                next_month = current_date.replace(day=28) + timedelta(days=4)
                next_month_start = next_month - timedelta(days=next_month.day - 1)

                monthly_payments = MembershipPayment.objects.filter(
                    payment_date__gte=current_date,
                    payment_date__lt=next_month_start
                )

                monthly_total = monthly_payments.aggregate(total=Sum('amount'))['total'] or 0
                month_totals[current_date.strftime('%Y-%m')] = monthly_total

                for payment in monthly_payments:
                    revenue_data.append([
                        payment.id,
                        payment.membership.id,
                        payment.amount,
                        payment.payment_date.strftime('%Y-%m-%d %H:%M:%S'),
                        payment.payment_method
                    ])

                current_date = next_month_start

            ws.append(['Payment ID', 'Membership ID', 'Amount', 'Payment Date', 'Payment Method'])
            for row in revenue_data:
                ws.append(row)

            ws.append([])
            ws.append(['Month', 'Total Revenue'])
            for month, total in month_totals.items():
                ws.append([month, total])

        wb.save(response)
        return response


class RoomRatingListView(ListView):
    model = RoomRating
    template_name = 'admins/room_ratings.html'
    context_object_name = 'ratings'
    paginate_by = 5
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['room'] = get_object_or_404(Room, pk=self.kwargs['pk'])
        return context